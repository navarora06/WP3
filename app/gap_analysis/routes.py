import os
from datetime import datetime

from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, send_file, abort, jsonify
from flask_login import login_required, current_user
from sqlalchemy import select, desc
from werkzeug.utils import secure_filename

from app.models import Interview, SupportDoc, GapReport, GapItem, Status, ACTION_OPTIONS
from app.util import db_session
from tasks.gap import gap_analysis_task

bp = Blueprint("gap", __name__, url_prefix="/gap")


@bp.get("/new")
@login_required
def new_gap():
    uid = int(current_user.get_id())

    with db_session() as db:
        interviews = (
            db.query(Interview)
            .filter(Interview.created_by == uid, Interview.status == Status.READY)
            .order_by(Interview.created_at.desc())
            .all()
        )

        docs = (
            db.query(SupportDoc)
            .filter(SupportDoc.created_by == uid, SupportDoc.status == Status.READY)
            .order_by(SupportDoc.created_at.desc())
            .all()
        )

        reports = (
            db.query(GapReport)
            .filter(GapReport.created_by == uid)
            .order_by(GapReport.created_at.desc())
            .all()
        )

        interview_map = {i.id: i.title for i in interviews}
        doc_map = {d.id: d.title for d in docs}
        for r in reports:
            if r.interview_id not in interview_map:
                iv = db.get(Interview, r.interview_id)
                if iv:
                    interview_map[iv.id] = iv.title
            if r.doc_id not in doc_map:
                dc = db.get(SupportDoc, r.doc_id)
                if dc:
                    doc_map[dc.id] = dc.title

    return render_template(
        "gap_analysis/new.html",
        interviews=interviews,
        docs=docs,
        reports=reports,
        interview_map=interview_map,
        doc_map=doc_map,
    )


@bp.post("/run")
@login_required
def run_gap():
    uid = int(current_user.get_id())
    interview_id = request.form.get("interview_id", type=int)
    doc_id = request.form.get("doc_id", type=int)

    if not interview_id or not doc_id:
        flash("Please select both an interview and a supporting document.", "danger")
        return redirect(url_for("gap.new_gap"))

    with db_session() as db:
        interview = db.get(Interview, interview_id)
        doc = db.get(SupportDoc, doc_id)

        if not interview or interview.created_by != uid or interview.status != Status.READY:
            flash("Invalid interview selection.", "danger")
            return redirect(url_for("gap.new_gap"))

        if not doc or doc.created_by != uid or doc.status != Status.READY:
            flash("Invalid document selection.", "danger")
            return redirect(url_for("gap.new_gap"))

        report = GapReport(
            interview_id=interview_id,
            doc_id=doc_id,
            created_by=uid,
            status=Status.PROCESSING,
        )
        db.add(report)
        db.flush()
        report_id = report.id

    gap_analysis_task.delay(report_id)

    flash("Gap analysis started. It will appear below when complete.", "success")
    return redirect(url_for("gap.new_gap"))


@bp.get("/report/<int:report_id>")
@login_required
def view_report(report_id: int):
    uid = int(current_user.get_id())
    with db_session() as db:
        rpt = db.get(GapReport, report_id)
        if not rpt:
            abort(404)
        if rpt.created_by != uid:
            abort(403)

        items = (
            db.query(GapItem)
            .filter(GapItem.report_id == report_id)
            .order_by(GapItem.id)
            .all()
        )

        rows = []
        for it in items:
            rows.append({
                "id": it.id,
                "claim": it.claim_text,
                "label": it.label.value if it.label else "UNKNOWN",
                "interview_evidence": it.interview_evidence or "",
                "doc_evidence": it.doc_evidence or "",
                "confidence": it.confidence or "Low",
                "suggestion": it.action_suggestion or "",
            })

        summary = rpt.summary_json or {}
        out_of_scope = (rpt.report_json or {}).get("out_of_scope", [])

    return render_template(
        "gap_analysis/report_view.html",
        rpt=rpt,
        rows=rows,
        summary=summary,
        out_of_scope=out_of_scope,
        action_options=ACTION_OPTIONS,
    )


@bp.post("/report/<int:report_id>/submit-review")
@login_required
def submit_review(report_id: int):
    """Batch-update all Action dropdown values from the view page."""
    uid = int(current_user.get_id())
    data = request.get_json(silent=True) or {}
    items_data = data.get("items", [])

    if not items_data:
        return jsonify({"error": "No items provided"}), 400

    action_set = set(ACTION_OPTIONS)
    with db_session() as db:
        rpt = db.get(GapReport, report_id)
        if not rpt or rpt.created_by != uid:
            return jsonify({"error": "Not found"}), 404

        for entry in items_data:
            item_id = entry.get("id")
            action = entry.get("action")
            if not item_id or action not in action_set:
                continue
            item = db.get(GapItem, item_id)
            if item and item.report_id == report_id:
                item.action_suggestion = action

        rpt.is_reviewed = True
        rpt.reviewed_at = datetime.utcnow()

    return jsonify({"ok": True})


@bp.get("/download/<int:report_id>")
@login_required
def download_report(report_id: int):
    uid = int(current_user.get_id())
    with db_session() as db:
        rpt = db.get(GapReport, report_id)
        if not rpt:
            abort(404)
        if rpt.created_by != uid:
            abort(403)
        key = rpt.report_storage_key
        fname = (rpt.report_name or f"gap_report_{report_id}") + ".xlsx"

    if not key:
        abort(404)

    path = current_app.storage.resolve_path(key)
    if not os.path.exists(path):
        abort(404)

    return send_file(
        path,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.post("/report/<int:report_id>/upload-reviewed")
@login_required
def upload_reviewed(report_id: int):
    """Accept a reviewed Excel report, validate format + action values, replace stored report."""
    uid = int(current_user.get_id())

    with db_session() as db:
        rpt = db.get(GapReport, report_id)
        if not rpt or rpt.created_by != uid:
            flash("Report not found.", "danger")
            return redirect(url_for("gap.new_gap"))
        if rpt.status != Status.READY:
            flash("Only READY reports can accept a reviewed upload.", "danger")
            return redirect(url_for("gap.new_gap"))
        stored_item_count = db.query(GapItem).filter(GapItem.report_id == report_id).count()

    f = request.files.get("reviewed_file")
    if not f or not f.filename.lower().endswith(".xlsx"):
        flash("Please upload a valid .xlsx file.", "danger")
        return redirect(url_for("gap.new_gap"))

    from openpyxl import load_workbook
    from io import BytesIO

    raw_bytes = f.read()
    try:
        wb = load_workbook(BytesIO(raw_bytes))
    except Exception:
        flash("Could not read the uploaded Excel file.", "danger")
        return redirect(url_for("gap.new_gap"))

    required_sheets = {"Gap_Analysis", "Out_of_Scope", "Summary"}
    if not required_sheets.issubset(set(wb.sheetnames)):
        flash(f"Missing sheets. Required: {', '.join(sorted(required_sheets))}.", "danger")
        return redirect(url_for("gap.new_gap"))

    ws = wb["Gap_Analysis"]
    from tasks.report_excel import GAP_HEADERS
    expected_headers = GAP_HEADERS
    actual_headers = [cell.value for cell in ws[1][:7]]
    if actual_headers != expected_headers:
        flash("Gap_Analysis sheet headers do not match the expected format.", "danger")
        return redirect(url_for("gap.new_gap"))

    ws_summary = wb["Summary"]
    embedded_report_id = None
    for row in ws_summary.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] == "Report ID":
            try:
                embedded_report_id = int(row[1])
            except (TypeError, ValueError):
                pass
            break

    if embedded_report_id is None or embedded_report_id != report_id:
        flash(f"Report ID mismatch: this file belongs to report {embedded_report_id}, not {report_id}.", "danger")
        return redirect(url_for("gap.new_gap"))

    action_set = set(ACTION_OPTIONS)
    reviewed_actions = {}
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=7, values_only=True), start=2):
        if row[0] is None:
            break
        action_val = (row[6] or "").strip() if row[6] else ""
        if action_val and action_val not in action_set:
            errors.append(f"Row {row_idx}: invalid Action '{action_val}'.")
        reviewed_actions[row_idx - 2] = action_val

    if errors:
        flash("Upload rejected. " + " ".join(errors[:5]), "danger")
        return redirect(url_for("gap.new_gap"))

    if len(reviewed_actions) != stored_item_count:
        flash(f"Row count mismatch: uploaded {len(reviewed_actions)} vs stored {stored_item_count} claims.", "danger")
        return redirect(url_for("gap.new_gap"))

    with db_session() as db:
        rpt = db.get(GapReport, report_id)
        items = (
            db.query(GapItem)
            .filter(GapItem.report_id == report_id)
            .order_by(GapItem.id)
            .all()
        )

        for idx, item in enumerate(items):
            action_val = reviewed_actions.get(idx, "")
            if action_val:
                item.action_suggestion = action_val

        safe = secure_filename(f.filename or "reviewed.xlsx")
        key = f"reports/gap_report_{report_id}_reviewed_{safe}"
        excel_path = current_app.storage.resolve_path(key)
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        with open(excel_path, "wb") as out:
            out.write(raw_bytes)

        rpt.report_storage_key = key
        rpt.is_reviewed = True
        rpt.reviewed_at = datetime.utcnow()

    flash("Reviewed report uploaded successfully.", "success")
    return redirect(url_for("gap.view_report", report_id=report_id))


@bp.post("/report/<int:report_id>/create-knowledge")
@login_required
def create_knowledge(report_id: int):
    """Trigger the knowledge creation pipeline for a reviewed gap report."""
    uid = int(current_user.get_id())

    with db_session() as db:
        rpt = db.get(GapReport, report_id)
        if not rpt or rpt.created_by != uid:
            flash("Report not found.", "danger")
            return redirect(url_for("gap.new_gap"))
        if not rpt.is_reviewed:
            flash("Report must be reviewed before creating knowledge.", "danger")
            return redirect(url_for("gap.view_report", report_id=report_id))

    from tasks.knowledge import create_knowledge_task
    create_knowledge_task.delay(report_id)

    flash("Knowledge creation started. This runs in the background.", "success")
    return redirect(url_for("gap.view_report", report_id=report_id))

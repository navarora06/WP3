"""
Generate a 3-tab Excel gap analysis report matching the project format:
  Tab 1 – Gap_Analysis   (claim table)
  Tab 2 – Out_of_Scope   (filtered sentences)
  Tab 3 – Summary         (metrics)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BODY_ALIGN = Alignment(vertical="top", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_LABEL_FILLS = {
    "SUPPORTED": PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid"),
    "CONTRADICTED": PatternFill(start_color="C62828", end_color="C62828", fill_type="solid"),
    "UNKNOWN": PatternFill(start_color="F9A825", end_color="F9A825", fill_type="solid"),
}
_LABEL_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)


def _style_header(ws, cols):
    for col_idx, name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


_ACTION_DEFAULTS = {
    "SUPPORTED": "Already documented",
    "CONTRADICTED": "Investigate Further",
    "UNKNOWN": "Confirm in next review",
}

_ACTION_OPTIONS = [
    "Already documented",
    "Confirm in next review",
    "Investigate Further",
    "Add to documentation",
    "Discard",
]


GAP_HEADERS = ["Claim", "Label", "Interview Evidence", "Doc Evidence",
               "Confidence", "Action Suggestion", "Action"]


def generate_gap_report_excel(output_path: str, data: dict, report_id: int,
                              report_name: str = ""):
    wb = Workbook()

    # ---- Tab 1: Gap_Analysis ----
    ws1 = wb.active
    ws1.title = "Gap_Analysis"

    _style_header(ws1, GAP_HEADERS)

    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 35
    ws1.column_dimensions["D"].width = 35
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 30
    ws1.column_dimensions["G"].width = 22

    action_dv = DataValidation(
        type="list",
        formula1='"' + ",".join(_ACTION_OPTIONS) + '"',
        allow_blank=False,
    )
    action_dv.error = "Please select a valid action."
    action_dv.errorTitle = "Invalid Action"
    action_dv.prompt = "Choose an action"
    action_dv.promptTitle = "Action"
    ws1.add_data_validation(action_dv)

    for row_idx, item in enumerate(data.get("gap_analysis", []), start=2):
        label = item.get("label", "UNKNOWN").upper()
        confidence = item.get("confidence", "Low")
        suggestion = item.get("action_suggestion") or _ACTION_DEFAULTS.get(label, "Confirm in next review")
        action_val = item.get("action", "") or ""

        ws1.cell(row=row_idx, column=1, value=item.get("claim", ""))
        label_cell = ws1.cell(row=row_idx, column=2, value=label)
        ws1.cell(row=row_idx, column=3, value=item.get("interview_evidence", ""))
        ws1.cell(row=row_idx, column=4, value=item.get("doc_evidence", ""))
        ws1.cell(row=row_idx, column=5, value=confidence)
        ws1.cell(row=row_idx, column=6, value=suggestion)
        action_cell = ws1.cell(row=row_idx, column=7, value=action_val)
        action_dv.add(action_cell)

        if label in _LABEL_FILLS:
            label_cell.fill = _LABEL_FILLS[label]
            label_cell.font = _LABEL_FONT
            label_cell.alignment = Alignment(horizontal="center", vertical="center")

        for c in range(1, 8):
            cell = ws1.cell(row=row_idx, column=c)
            cell.border = _THIN_BORDER
            if c != 2:
                cell.alignment = _BODY_ALIGN

    # ---- Tab 2: Out_of_Scope ----
    ws2 = wb.create_sheet("Out_of_Scope")

    _style_header(ws2, ["Filtered Sentence (out-of-scope / not a claim)"])
    ws2.column_dimensions["A"].width = 100

    for row_idx, item in enumerate(data.get("out_of_scope", []), start=2):
        sentence = item if isinstance(item, str) else item.get("sentence", str(item))
        cell = ws2.cell(row=row_idx, column=1, value=sentence)
        cell.alignment = _BODY_ALIGN
        cell.border = _THIN_BORDER

    # ---- Tab 3: Summary ----
    ws3 = wb.create_sheet("Summary")

    _style_header(ws3, ["Metric", "Value"])
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 15

    summary = data.get("summary", {})
    metrics = [
        ("Report ID", report_id),
        ("Report Name", report_name),
        ("Total claims included", summary.get("total_claims", 0)),
        ("Supported", summary.get("supported", 0)),
        ("Contradicted", summary.get("contradicted", 0)),
        ("Unknown", summary.get("unknown", 0)),
        ("Out-of-scope filtered", summary.get("out_of_scope_filtered", 0)),
        ("Doc blocks after dedupe", summary.get("doc_blocks_after_dedupe", 0)),
    ]

    for row_idx, (metric, value) in enumerate(metrics, start=2):
        m_cell = ws3.cell(row=row_idx, column=1, value=metric)
        v_cell = ws3.cell(row=row_idx, column=2, value=value)
        m_cell.border = _THIN_BORDER
        v_cell.border = _THIN_BORDER
        m_cell.alignment = _BODY_ALIGN
        v_cell.alignment = Alignment(horizontal="right", vertical="top")

    wb.save(output_path)
